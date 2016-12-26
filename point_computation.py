from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font


# ------------------------------------------ FUNCTIONS --------------------------------------------


# Returns a list with the team's latest 6 game results at the time. Input is team name and the row which it should look backwards of.
def look_back(team, row):

	for rev_row in range(row-1, 1,-1):						# Look at previous rows to find team name

			for rev_col in range(3,5):

				if team == ws.cell(row = rev_row, column = rev_col).value:		# If team is found

					if rev_col == 3:											# Check if it was Home or Away

						played = "H"
					else:

						played = "A"

					team_history = eval(ws.cell(row = rev_row, column = rev_col + 8).value)															# Read team_history from the team's latest game
					
					team_history = [(played, int(ws.cell(row = rev_row, column = 5).value), int(ws.cell(row = rev_row, column = 6).value))] + team_history		# Append latest game to the front of the list e.g. (H, 2, 1)
					
					if len(team_history) > 6:																										# List deletes the oldest game whenever it exceeds a history of 6 games.

						del team_history[-1]

					return team_history

	return []																								# If no match is found (for the first games), return []




# ------------------------------------------ START --------------------------------------------

wb_final = Workbook()	#NEW

ws_final = wb_final.active			#NEW

row_final = 2					#NEW


for z in range(2005, 2012):

	wb = load_workbook("all-euro-data-%d-%d-FIXED.xlsx" % (z, z+1))

	for ws in wb:

		ws.cell(row= 1, column= 13, value= "Points").font = Font(bold=True)		# Bold column title
		ws.cell(row= 1, column= 12, value= "Away History").font = Font(bold=True)		# Bold column title
		ws.cell(row= 1, column= 11, value= "Home History").font = Font(bold=True)		# Bold column title

		# --- Start gathering last 6 games' history ---

		for row in range(2,ws.max_row+1):			# ws.max_row+1 Exclude title AND fist match, which will raise error when trying to look backwards of it.
			
			for col in range(3,5):					# Home, Away
				
				print(ws, row, col)
					
				team = ws.cell(row = row, column = col).value
				
				print(team)
				
				team_history = look_back(team, row)

				ws.cell(row= row, column= col + 8 , value= str(team_history))					# Append latest game to list e.g. (H, 2, 1)


		# --- Start calculating points ---

				if len(team_history) == 6:

					try:
						points = float(ws.cell(row = row, column = 13).value)					# Grabs points that home history generated
					
					except TypeError:															# If points cell is None, add sligtly/heavy favorite points which have to be added only once. In this case the fist time.
						points = 0

						try:
							home_odds = float(ws.cell(row = row, column = 7).value)

							away_odds = float(ws.cell(row = row, column = 8).value)

						except TypeError:												# if 1x2 odds empty, paint points blue

							home_odds = 2.5
							away_odds = 1.5
							ws.cell(row= row, column= 13).font = Font(bold=True,color='0000ff')		# Blue bold text


						if home_odds < 2:					# Home heavy favorite

							points -= 2

						elif away_odds < 2:					# Away heavy favorite

							points -= 1

						elif home_odds > away_odds:			# Away slightly favorite

							points += 2

						else:								# Home slightly favorite

							points += 1

					for game in team_history:

						if all(g > 0 for g in game[-2:]):												# for each btts give one point

							if (col == 3 and game[0] == "H") or (col == 4 and game[0] == "A"):				# extra point if btts had home as home or away as away

								points += 1

							points += 1

						for i in range(1,3):										# +0.5 pts for each goal scored or conceded over 2

							if game[i] > 2:

								points += (game[i] - 2) * 0.5

						if all(g == 0 for g in game[-2:]):							# -2 pts for 0:0 score

							points -= 2

					print(points)		
					
					ws.cell(row= row, column= 13 , value= points)					# Append latest game to list e.g. (H, 2, 1)
#NEW

			if ws.cell(row= row, column= 13).value != None:

				for col in range(1,14):

					ws_final.cell(row = row_final, column = col, value = ws.cell(row= row, column= col).value)

				row_final += 1



for col in range(1,14):

	ws_final.cell(row= 1, column= col, value = ws.cell(row= 1, column= col).value)

wb_final.save("all-euro-data-2005-2012-FINAL.xlsx")


#	------ Goal sum, Line, line odds, LineP/L -----------


wb = load_workbook("all-euro-data-2005-2012-FINAL.xlsx", data_only=True)			# data_only=True avoids reading the formulas from inside the cells

ws = wb.active

conversion = ((1.25, 4, 1.9), (1.275, 3.75, 1.9), (1.299, 3.75, 2), 			# Data to convert Over 2.5 odds to Line & Line odds
	(1.3, 3.5, 1.8), (1.325, 3.5, 1.85), (1.35, 3.5, 1.9), 
	(1.375, 3.5, 1.95), (1.399, 3.5, 2), (1.4, 3.25, 1.8), 
	(1.425, 3.25, 1.85), (1.45, 3.25, 1.9), (1.475, 3.25, 1.95), 
	(1.499, 3.25, 2), (1.5, 3, 1.8), (1.525, 3, 1.85), (1.55, 3, 1.9), 
	(1.575, 3, 1.95), (1.599, 3, 2), (1.6, 2.75, 1.8), (1.65, 2.75, 1.85), 
	(1.7, 2.75, 1.9), (1.75, 2.75, 1.95), (1.799, 2.75, 2), (1.8, 2.5, 1.8), 
	(1.85, 2.5, 1.85), (1.9, 2.5, 1.9), (1.95, 2.5, 1.95), (1.99, 2.5, 1.99), 
	(2.00, 2.25, 1.8), (2.05, 2.25, 1.85), (2.1, 2.25, 1.9), (2.15, 2.25, 1.95), 
	(2.2, 2.25, 2), (2.25, 2.25, 2.025), (2.40, 2, 1.9), (2.5, 2, 1.925), (2.6, 2, 1.975), 
	(2.7, 2, 2), (2.8, 1.75, 1.85), (2.9, 1.75, 1.9), (3.1, 1.75, 1.95), (4, 1.5, 1.8))


ws.cell(row= 1, column= 14, value= "Goals")
ws.cell(row= 1, column= 15, value= "Line")
ws.cell(row= 1, column= 16, value= "Line Odds")
ws.cell(row= 1, column= 17, value= "Line P/L")

for row in range(2, ws.max_row):

	o_25 = ws.cell(row = row, column = 9).value									# Calculates line and line odds according to the tuple above.

	goal_sum = ws.cell(row = row, column = 5).value	+ ws.cell(row = row, column = 6).value

	ws.cell(row = row, column = 14, value = goal_sum)									# Writes sum of goals

	for item in conversion:

		if o_25 <= item[0]:

			line = item[1]

			ws.cell(row = row, column = 15, value = line)

			line_odds = item[2]

			ws.cell(row = row, column = 16, value = line_odds)

			break

	goals = ws.cell(row = row, column = 14).value									# Writes line and line odds under respective column.

	diff = goals - line
	
	if diff > 0.25:																	# Calculates Profit/Loss from Goals, line and line odds. Writes it.

		ws.cell(row = row, column = 17, value = 10*(line_odds-1))

	elif diff == 0.25:

		ws.cell(row = row, column = 17, value = 5*(line_odds-1))

	elif diff == 0:

		ws.cell(row = row, column = 17, value = 0)

	elif diff == -0.25:

		ws.cell(row = row, column = 17, value = -5*(line_odds-1))

	else:

		ws.cell(row = row, column = 17, value = -10)

	print(o_25, line, line_odds, ws.cell(row = row, column = 17).value)


wb.save("UBER_SHEET_2005-2012_data_FINAL.xlsx")
